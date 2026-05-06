import io
import struct
import zlib
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET


def extract_text(filename: str, content: bytes) -> str:
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        return _parse_pdf(content)
    elif ext in (".docx",):
        return _parse_docx(content)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(content)
    elif ext == ".hwp":
        return _parse_hwp(content)
    elif ext == ".hwpx":
        return _parse_hwpx(content)
    elif ext in (".csv",):
        return content.decode("utf-8", errors="replace")
    else:
        return content.decode("utf-8", errors="replace")


def _parse_pdf(content: bytes) -> str:
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(content))
    texts = []
    for page in reader.pages:
        try:
            text = page.extract_text()
            if text:
                texts.append(text)
        except Exception:
            continue
    return "\n".join(texts)


def _parse_docx(content: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _parse_excel(content: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"[시트: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append("\t".join(cells))
    return "\n".join(lines)


def _parse_hwp(content: bytes) -> str:
    import olefile

    ole = olefile.OleFileIO(io.BytesIO(content))
    try:
        header = ole.openstream("FileHeader").read()
        is_compressed = bool(header[36] & 1)

        texts = []
        for entry in ole.listdir():
            if entry[0] == "BodyText":
                data = ole.openstream(entry).read()
                if is_compressed:
                    try:
                        data = zlib.decompress(data, -15)
                    except zlib.error:
                        continue
                text = _extract_text_from_hwp_section(data)
                if text:
                    texts.append(text)

        return "\n".join(texts) if texts else ""
    finally:
        ole.close()


def _extract_text_from_hwp_section(data: bytes) -> str:
    texts = []
    offset = 0
    while offset < len(data):
        if offset + 4 > len(data):
            break
        header = struct.unpack_from("<I", data, offset)[0]
        tag_id = header & 0x3FF
        level = (header >> 10) & 0x3FF
        size = (header >> 20) & 0xFFF

        offset += 4
        if size == 0xFFF:
            if offset + 4 > len(data):
                break
            size = struct.unpack_from("<I", data, offset)[0]
            offset += 4

        if offset + size > len(data):
            break

        # HWPTAG_PARA_TEXT = 67
        if tag_id == 67:
            text = _decode_hwp_para_text(data[offset:offset + size])
            if text:
                texts.append(text)

        offset += size

    return "\n".join(texts)


def _decode_hwp_para_text(data: bytes) -> str:
    chars = []
    i = 0
    while i + 1 < len(data):
        code = struct.unpack_from("<H", data, i)[0]
        i += 2
        if code == 0:
            break
        # Control characters in HWP: skip inline controls
        if code < 32:
            if code in (1, 2, 3, 11, 12, 14, 15, 16, 17, 18, 21, 22, 23):
                # Extended control: skip additional 14 bytes
                i += 14
            elif code == 10:
                chars.append("\n")
            elif code == 13:
                chars.append("\n")
            # Other control chars: skip
        else:
            chars.append(chr(code))
    return "".join(chars).strip()


def _parse_hwpx(content: bytes) -> str:
    texts = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            section_files = sorted(
                [n for n in zf.namelist() if "section" in n.lower() and n.endswith(".xml")]
            )
            for section_file in section_files:
                xml_data = zf.read(section_file)
                text = _extract_text_from_hwpx_xml(xml_data)
                if text:
                    texts.append(text)
    except zipfile.BadZipFile:
        return ""
    return "\n".join(texts)


def _extract_text_from_hwpx_xml(xml_data: bytes) -> str:
    texts = []
    try:
        root = ET.fromstring(xml_data)
        # HWPX uses namespaces; find all text elements regardless of namespace
        for elem in root.iter():
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if tag == "t" and elem.text:
                texts.append(elem.text)
            elif tag == "p" and texts and texts[-1] != "\n":
                texts.append("\n")
    except ET.ParseError:
        return ""
    return "".join(texts).strip()
